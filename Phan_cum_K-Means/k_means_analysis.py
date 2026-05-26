import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

# ─── THIẾT LẬP ĐƯỜNG DẪN HỆ THỐNG CỐ ĐỊNH ───────────────────────────────────
FILE_INPUT = 'student_grades_input.xlsx'
CHART_OUTPUT = 'gpa_distribution_chart.png'
EXCEL_OUTPUT = 'KMeans_Student_Segmentation_K58KTP.xlsx'

if not os.path.exists(FILE_INPUT):
    if os.path.exists('data.xlsx'):
        FILE_INPUT = 'data.xlsx'
    else:
        raise FileNotFoundError(f"Không tìm thấy tập dữ liệu cấu trúc đầu vào.")

# ─── ĐỊNH NGHĨA CÁC HÀM BỔ TRỢ (HELPER FUNCTIONS) ───────────────────────────
def chuan_hoa_diem(v):
    if pd.isna(v): return np.nan
    s = str(v).strip().lower().replace('nan', '')
    if s in ['-', '—', 'x', '']: return np.nan
    try:
        f = float(s)
        return f if 0 <= f <= 4.0 else np.nan
    except: 
        return np.nan

def phan_tich_mon_hoc(row_data):
    valid_grades = row_data.dropna()
    total_subjects = len(valid_grades)
    if total_subjects == 0: return "Không có dữ liệu điểm"
    num_A = sum(valid_grades >= 3.5)
    num_F = sum(valid_grades < 1.0)
    if num_F > 0: return f"Còn nợ {int(num_F)} môn chưa đạt (F)"
    elif num_A >= (total_subjects * 0.4): return f"Vượt trội ({int(num_A)} môn đạt điểm A)"
    else: return "Phong độ học tập ổn định"

# ĐÃ SỬA: Đưa hàm này lên trên trước khi gọi ở phần dựng Tab 3
def thiet_lap_cum_cot(ws, start_col, ten_nhom, ma_mau):
    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col+3)
    top_cell = ws.cell(row=3, column=start_col, value=ten_nhom)
    top_cell.font = Font(name='Segoe UI', size=10, bold=True, color='1F4E78')
    top_cell.fill = PatternFill('solid', start_color=ma_mau, end_color=ma_mau)
    top_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    tieu_de_con = ['STT', 'Họ và Tên', 'GPA', 'Đánh giá môn']
    for i, txt in enumerate(tieu_de_con):
        cell = ws.cell(row=4, column=start_col+i, value=txt)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

# ─── 1. TIỀN XỬ LÝ VÀ TRÍCH XUẤT ĐẶC TRƯNG (FEATURE ENGINEERING) ───────────
df_raw = pd.read_excel(FILE_INPUT, header=None)

mssv_list = df_raw.iloc[1, 3:].tolist()
name_list = df_raw.iloc[2, 3:].tolist()
grade_data = df_raw.iloc[4:, 3:].copy().T  # Ma trận chuyển vị (Students x Features)

# Chuẩn hóa ma trận điểm số chi tiết
grade_cleaned = grade_data.map(chuan_hoa_diem)

# KỸ THUẬT PHÁI SINH ĐẶC TRƯNG (FEATURE ENGINEERING)
features_engineered = pd.DataFrame()
features_engineered['Nu_Mon_F'] = grade_cleaned.apply(lambda row: sum(row < 1.0), axis=1)
features_engineered['Ty_Le_Mon_A'] = grade_cleaned.apply(lambda row: sum(row >= 3.5) / len(row.dropna()) if not row.dropna().empty else 0, axis=1)

# Xử lý giá trị khuyết (Mean Imputation) theo từng môn học
grade_imputed = grade_cleaned.apply(lambda col: col.fillna(col.mean() if not col.dropna().empty else 0), axis=0)

# Hợp nhất ma trận điểm ban đầu và các đặc trưng chỉ báo hành vi mới
X_matrix = pd.concat([grade_imputed, features_engineered], axis=1)
gpa_series = grade_cleaned.mean(axis=1).round(2)

danh_gia_list = [phan_tich_mon_hoc(grade_cleaned.iloc[i]) for i in range(len(grade_cleaned))]

# ─── 2. THỰC THI MÔ HÌNH HỌC MÁY KHÔNG GIÁM SÁT (K-MEANS CLUSTERING) ────────
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X_matrix.values)

# Thực hiện phân cụm K-Means đa chiều (Giữ nguyên không ép chiều dữ liệu)
kmeans = KMeans(n_clusters=3, init='k-means++', random_state=42, n_init=10)
cluster_labels = kmeans.fit_predict(X_scaled)

df_analytics = pd.DataFrame({
    'MSSV': mssv_list,
    'Ho_va_ten': name_list,
    'GPA': gpa_series.values,
    'Danh_gia': danh_gia_list,
    'Cluster_Raw': cluster_labels
})
df_analytics = df_analytics.dropna(subset=['GPA']).reset_index(drop=True)

# Sắp xếp tên cụm theo giá trị GPA trung bình giảm dần
cluster_gpa_means = df_analytics.groupby('Cluster_Raw')['GPA'].mean().sort_values(ascending=False)
cluster_mapping = {
    cluster_gpa_means.index[0]: 'Cluster_High_Achievers',     
    cluster_gpa_means.index[1]: 'Cluster_Steady_Performers',   
    cluster_gpa_means.index[2]: 'Cluster_At_Risk'              
}
df_analytics['Xep_loai'] = df_analytics['Cluster_Raw'].map(cluster_mapping)

# Sắp xếp bảng tổng thể theo thứ tự điểm số giảm dần để dễ theo dõi
df_sorted = df_analytics.sort_values(by='GPA', ascending=False).reset_index(drop=True)
df_gioi = df_sorted[df_sorted['Xep_loai'] == 'Cluster_High_Achievers'].reset_index(drop=True)
df_kha = df_sorted[df_sorted['Xep_loai'] == 'Cluster_Steady_Performers'].reset_index(drop=True)
df_tb = df_sorted[df_sorted['Xep_loai'] == 'Cluster_At_Risk'].reset_index(drop=True)
total_count = len(df_sorted)

# ─── 3. TRỰC QUAN HÓA KẾT QUẢ ĐỒ HỌA CHUYÊN NGHIỆP (MATPLOTLIB) ─────────────
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Segoe UI', 'Calibri', 'Arial']

counts = [len(df_gioi), len(df_kha), len(df_tb)]
labels = [
    f'High Achievers\n(Phân khối Xuất sắc)\nN = {len(df_gioi)}', 
    f'Steady Performers\n(Phân khối Ổn định)\nN = {len(df_kha)}', 
    f'At-Risk\n(Phân khối Cần chú ý)\nN = {len(df_tb)}'
]
colors = ['#A9D08E', '#FFE699', '#F4B084']

fig, ax = plt.subplots(figsize=(8, 5))
bars = ax.bar(labels, counts, color=colors, edgecolor='#BFBFBF', width=0.45, zorder=3)

ax.set_title('MÔ HÌNH PHÂN PHỐI TẦN SUẤT KHỐI SINH VIÊN QUA THUẬT TOÁN K-MEANS', fontsize=12, fontweight='bold', pad=22, color='#1F4E78')
ax.set_ylabel('Số lượng sinh viên (Mẫu N)', fontsize=10, fontweight='bold', color='#404040')
ax.grid(axis='y', linestyle='--', alpha=0.4, zorder=0)

for bar in bars:
    yval = bar.get_height()
    ax.text(bar.get_x() + bar.get_width()/2, yval + 0.4, f"{int(yval)} SV\n({yval/total_count*100:.1f}%)", 
            ha='center', va='bottom', fontsize=9, fontweight='bold', color='#262626')

ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.spines['left'].set_color('#D9D9D9')
ax.spines['bottom'].set_color('#D9D9D9')
ax.tick_params(axis='both', colors='#404040', labelsize=9.5)

plt.tight_layout()
plt.savefig(CHART_OUTPUT, dpi=150)
plt.close()

# ─── 4. KIẾN TRÚC TẬP TIN EXCEL ĐA NHIỆM CHUẨN SONG NGỮ (OPENPYXL) ──────────
wb = Workbook()
COLOR_MAP = {'Cluster_High_Achievers': 'E2EFDA', 'Cluster_Steady_Performers': 'FFF2CC', 'Cluster_At_Risk': 'FCE4D6'}
HEADER_FILL = PatternFill('solid', start_color='2F5597', end_color='2F5597')
HEADER_FONT = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Segoe UI', size=13, bold=True, color='1F4E78')
TEXT_FONT = Font(name='Segoe UI', size=10, color='000000')
BOLD_FONT = Font(name='Segoe UI', size=10, bold=True, color='000000')
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

# --- TAB 1: Cấu trúc phân khối ---
ws1 = wb.active; ws1.title = 'Phân khối cấu trúc năng lực'
ws1.views.sheetView[0].showGridLines = True
ws1.merge_cells('A1:F1')
ws1['A1'] = 'BẢNG PHÂN TÍCH CẤU TRÚC PHÂN CỤM NĂNG LỰC SINH VIÊN TOÀN LỚP K58KTP'
ws1['A1'].font = TITLE_FONT; ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

headers = ['STT', 'MSSV', 'Họ và Tên', 'GPA Hệ 4', 'Nhãn Phân Cụm Thuật Toán (Segment)', 'Đánh giá cấu trúc môn học']
for col_idx, text in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col_idx, value=text)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws1.row_dimensions[3].height = 28

for idx, row in df_sorted.iterrows():
    r_idx = idx + 4
    xl = row['Xep_loai']
    fill_color = PatternFill('solid', start_color=COLOR_MAP[xl], end_color=COLOR_MAP[xl])
    
    vn_label = "Phân khối Xuất sắc" if xl == 'Cluster_High_Achievers' else ("Phân khối Ổn định" if xl == 'Cluster_Steady_Performers' else "Phân khối Cần chú ý")
    full_label = f"{xl} ({vn_label})"
    
    vals = [idx + 1, row['MSSV'], row['Ho_va_ten'], row['GPA'], full_label, row['Danh_gia']]
    for c_idx, val in enumerate(vals, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=val)
        cell.font = TEXT_FONT; cell.border = thin_border; cell.fill = fill_color
        cell.alignment = Alignment(horizontal='center' if c_idx in [1, 2, 4, 5] else 'left', vertical='center')
        if c_idx == 4: cell.number_format = '0.00'
    ws1.row_dimensions[r_idx].height = 20
widths1 = [6, 16, 28, 12, 36, 32]
for c_idx, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(c_idx)].width = w


# --- TAB 2: Trọng tâm phân phối hình học ---
ws2 = wb.create_sheet(title='Trọng tâm phân phối hình học')
ws2.views.sheetView[0].showGridLines = True
ws2.merge_cells('A1:D1')
ws2['A1'] = 'BÁO CÁO GIÁ TRỊ TRỌNG TÂM BIÊN ĐỘ PHÂN CỤM K-MEANS'
ws2['A1'].font = TITLE_FONT; ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 40

headers_tk = ['Phân cụm Thuật toán (Segment)', 'Đặc trưng GPA cụm (Mean ± Std)', 'Mẫu khảo sát (N)', 'Tỷ lệ mẫu (%)']
for col_idx, text in enumerate(headers_tk, 1):
    cell = ws2.cell(row=3, column=col_idx, value=text)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws2.row_dimensions[3].height = 28

tk_rows_data = [
    ('Cluster_High_Achievers (Phân khối Xuất sắc)', f"GPA trung bình: {df_gioi['GPA'].mean():.2f} (±{df_gioi['GPA'].std():.2f})", len(df_gioi), len(df_gioi)/total_count),
    ('Cluster_Steady_Performers (Phân khối Ổn định)', f"GPA trung bình: {df_kha['GPA'].mean():.2f} (±{df_kha['GPA'].std():.2f})", len(df_kha), len(df_kha)/total_count),
    ('Cluster_At_Risk (Phân khối Cần chú ý)', f"GPA trung bình: {df_tb['GPA'].mean():.2f} (±{df_tb['GPA'].std():.2f})", len(df_tb), len(df_tb)/total_count)
]

for idx, (nhom, tchi, count, pct) in enumerate(tk_rows_data):
    r_idx = idx + 4
    raw_key = nhom.split(' ')[0]
    fill_color = PatternFill('solid', start_color=COLOR_MAP[raw_key], end_color=COLOR_MAP[raw_key])
    c1 = ws2.cell(row=r_idx, column=1, value=nhom)
    c2 = ws2.cell(row=r_idx, column=2, value=tchi)
    c3 = ws2.cell(row=r_idx, column=3, value=count)
    c4 = ws2.cell(row=r_idx, column=4, value=pct)
    for cell in [c1, c2, c3, c4]:
        cell.font = TEXT_FONT; cell.border = thin_border; cell.fill = fill_color
        cell.alignment = Alignment(horizontal='center' if cell in [c3, c4] else 'left', vertical='center')
    c4.number_format = '0.0%'
    ws2.row_dimensions[r_idx].height = 24

r_total = 7
ws2.cell(row=r_total, column=1, value='Tổng cộng mẫu thực nghiệm')
ws2.cell(row=r_total, column=2, value='-')
ws2.cell(row=r_total, column=3, value=total_count)
ws2.cell(row=r_total, column=4, value=1.0)
for c_idx in range(1, 5):
    cell = ws2.cell(row=r_total, column=c_idx)
    cell.font = BOLD_FONT; cell.border = thin_border; cell.alignment = Alignment(horizontal='center', vertical='center')
ws2.cell(row=r_total, column=4).number_format = '0.0%'
ws2.row_dimensions[r_total].height = 24
widths2 = [42, 35, 18, 18]
for c_idx, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(c_idx)].width = w

if os.path.exists(CHART_OUTPUT):
    img = OpenpyxlImage(CHART_OUTPUT)
    ws2.add_image(img, 'A10')


# --- TAB 3: Phân khối song song ---
ws3 = wb.create_sheet(title='Cấu trúc khối phân cụm')
ws3.views.sheetView[0].showGridLines = True
ws3.merge_cells('A1:N1')
ws3['A1'] = 'DANH SÁCH CHI TIẾT THEO TỪNG PHÂN KHỐI ĐỐI TƯỢNG TỰ NHIÊN'
ws3['A1'].font = TITLE_FONT; ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 40

thiet_lap_cum_cot(ws3, start_col=1, ten_nhom='🏆 CLUSTER: HIGH ACHIEVERS (XUẤT SẮC)', ma_mau='E2EFDA')
thiet_lap_cum_cot(ws3, start_col=6, ten_nhom='✅ CLUSTER: STEADY PERFORMERS (ỔN ĐỊNH)', ma_mau='FFF2CC')
thiet_lap_cum_cot(ws3, start_col=11, ten_nhom='⚠️ CLUSTER: AT-RISK (CẦN CHÚ Ý)', ma_mau='FCE4D6')
ws3.row_dimensions[3].height = 25
ws3.row_dimensions[4].height = 22

max_rows = max(len(df_gioi), len(df_kha), len(df_tb))
for idx in range(max_rows):
    r_idx = idx + 5
    ws3.row_dimensions[r_idx].height = 20
    
    if idx < len(df_gioi):
        row = df_gioi.iloc[idx]
        fill = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
        for offset, val in enumerate([idx+1, row['Ho_va_ten'], row['GPA'], row['Danh_gia']]):
            cell = ws3.cell(row=r_idx, column=1+offset, value=val)
            cell.font = TEXT_FONT; cell.border = thin_border; cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if offset in [0, 2] else 'left', vertical='center')
            if offset == 2: cell.number_format = '0.00'
            
    if idx < len(df_kha):
        row = df_kha.iloc[idx]
        fill = PatternFill('solid', start_color='FFF2CC', end_color='FFF2CC')
        for offset, val in enumerate([idx+1, row['Ho_va_ten'], row['GPA'], row['Danh_gia']]):
            cell = ws3.cell(row=r_idx, column=6+offset, value=val)
            cell.font = TEXT_FONT; cell.border = thin_border; cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if offset in [0, 2] else 'left', vertical='center')
            if offset == 2: cell.number_format = '0.00'
            
    if idx < len(df_tb):
        row = df_tb.iloc[idx]
        fill = PatternFill('solid', start_color='FCE4D6', end_color='FCE4D6')
        for offset, val in enumerate([idx+1, row['Ho_va_ten'], row['GPA'], row['Danh_gia']]):
            cell = ws3.cell(row=r_idx, column=11+offset, value=val)
            cell.font = TEXT_FONT; cell.border = thin_border; cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if offset in [0, 2] else 'left', vertical='center')
            if offset == 2: cell.number_format = '0.00'

col_widths_s3 = {1: 5, 2: 24, 3: 8, 4: 26, 5: 4, 6: 5, 7: 24, 8: 8, 9: 26, 10: 4, 11: 5, 12: 24, 13: 8, 14: 26}
for col_num, width in col_widths_s3.items():
    ws3.column_dimensions[get_column_letter(col_num)].width = width

# ─── 5. LƯU TẬP TIN KẾT QUẢ KHOA HỌC DỮ LIỆU ─────────────────────────────────
wb.save(EXCEL_OUTPUT)

print("=" * 90)
print(f" PIPELINE PHÂN TÍCH K-MEANS HOÀN THÀNH THEO TIÊU CHUẨN KHOA HỌC DỮ LIỆU!")
print(f" Thư mục làm việc: ./Phan_cum_K-Means")
print(f" ── MÃ NGUỒN (Script):      {os.path.basename(__file__)}")
print(f" ── ĐỒ THỊ CHUẨN HÓA (Plot): {CHART_OUTPUT}")
print(f" ── BÁO CÁO CẤU TRÚC (Data): {EXCEL_OUTPUT}")
print("=" * 90)